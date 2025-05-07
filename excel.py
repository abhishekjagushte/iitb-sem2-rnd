import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# Path to the 'res' folder
res_path = 'res'

# Create a new Excel workbook
wb = Workbook()
wb.remove(wb.active)  # Remove the default sheet

# Iterate over subfolders in 'res'
for folder_name in os.listdir(res_path):
    folder_path = os.path.join(res_path, folder_name)
    if not os.path.isdir(folder_path):
        continue

    # Create a new sheet with the folder name
    ws = wb.create_sheet(title=folder_name)
    current_row = 1

    # Get sorted list of CSV files
    csv_files = sorted(f for f in os.listdir(folder_path) if f.endswith('.csv'))

    for csv_file in csv_files:
        csv_path = os.path.join(folder_path, csv_file)
        df = pd.read_csv(csv_path)

        # Add title without '.csv' and make it bold
        title = os.path.splitext(csv_file)[0]
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True)
        current_row += 1

        # Write the dataframe
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row + row_idx, column=col_idx, value=value)

                # Bold the header row
                if row_idx == 0:
                    cell.font = Font(bold=True)

                # Bold the first column
                if col_idx == 1:
                    cell.font = Font(bold=True)

                # Apply border to all cells
                cell.border = thin_border

        # Update row counter (df rows + header + 3-row gap)
        current_row += len(df) + 1 + 3

# Save workbook
wb.save("res.xlsx")
